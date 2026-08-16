"""
Master Financial Statement Extraction Pipeline
===============================================
Walks a folder of the shape:

    <root_folder>/
        AAPL/
            AAPL_2018_annual_report.pdf     <- skipped (pre-2022)
            AAPL_2023_10K.pdf                <- processed
            AAPL_2024_10K.pdf                <- processed
        MTN/
            MTN_2025_AFS_....pdf             <- processed
            ...

...finds every PDF whose filename indicates FY2022 or later, uses a cheap
classifier-model pass over filenames to keep only actual financial-statement
filings per ticker (AFS/IAR-type documents - dropping AGM notices, debt
prospectuses, remuneration reports, etc.), selecting exactly one filing per
distinct year found for each ticker (min_year through the most recent year
on disk), extracts the core financial statements (income statement, balance
sheet, cash flow statement) using Claude, and writes one combined Excel
workbook covering every ticker/year found.

    ┌─────────────────────────────────────────────────────────────────┐
    │  You'll be prompted for your Claude API key the first time this  │
    │  script actually needs to call the API - see get_client() below  │
    │  It is never written to disk or echoed to the terminal.          │
    └─────────────────────────────────────────────────────────────────┘

Usage (from a terminal):
    python financial_extraction_master.py /path/to/root_folder
    python financial_extraction_master.py /path/to/root_folder --min-year 2022
    python financial_extraction_master.py /path/to/root_folder --out results.xlsx

Usage (inside a Jupyter notebook cell — do NOT call main(), it uses argparse
which reads sys.argv and will crash with SystemExit: 2 in a notebook kernel):
    rows = run_extraction("/path/to/root_folder", min_year=2022)
"""

import argparse
import base64
import csv
import io
import json
import os
import re
import sys
from dataclasses import dataclass
from getpass import getpass
from pathlib import Path
from typing import Optional

import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# =============================================================================
# >>> API KEY <<<  ---------------------------------------------------------
# =============================================================================
# No key is hardcoded here on purpose. get_client() checks, in order:
#   1. the ANTHROPIC_API_KEY environment variable
#   2. an interactive getpass prompt (masked input, nothing echoed/saved)
# Get a key at https://platform.claude.com  (Anthropic's developer console).
MODEL_NAME = "claude-sonnet-5"            # extraction model
CLASSIFIER_MODEL_NAME = "claude-haiku-4-5-20251001"  # cheap model for the "does this page matter" pass
# =============================================================================

# `anthropic` is imported lazily (inside get_client(), not at the top of the
# file) so that folder-scanning/filtering can be run and tested without the
# SDK installed yet - only the actual extraction step needs it.


def get_client():
    import anthropic

    key = os.environ.get("ANTHROPIC_API_KEY")
    if not key:
        key = getpass("Enter your Anthropic API key: ")
        if not key:
            sys.exit("No API key entered - aborting.")
        os.environ["ANTHROPIC_API_KEY"] = key  # so later calls in this run don't re-prompt

    return anthropic.Anthropic(api_key=key)


# =============================================================================
# FOLDER DISCOVERY + YEAR FILTERING
# =============================================================================

YEAR_PATTERN = re.compile(r"(20\d{2})")


def extract_year_from_filename(filename: str) -> Optional[int]:
    """Pull every 4-digit 20xx year out of the filename and return the most
    recent one (filings often mention both a filing year and a prior-year
    comparative, e.g. 'fy25... 2025-signed.pdf' vs 'notice-of-agm-2018.pdf').
    Returns None if no year is found in the name at all."""
    years = [int(y) for y in YEAR_PATTERN.findall(filename)]
    return max(years) if years else None


@dataclass
class DiscoveredFile:
    ticker: str
    path: Path
    year: Optional[int]


def discover_documents(root_folder: Path, min_year: int, include_undated: bool = True) -> list[DiscoveredFile]:
    """Ticker = immediate subfolder name under root_folder. Only *.pdf files
    are considered. Files whose filename-derived year is < min_year are
    skipped. Files with no year in the name are included by default (with a
    printed warning) since we can't tell their vintage from the name alone -
    set include_undated=False to skip them instead."""
    found = []
    skipped = []
    for ticker_dir in sorted(p for p in root_folder.iterdir() if p.is_dir()):
        ticker = ticker_dir.name
        for pdf_path in sorted(ticker_dir.rglob("*.pdf")):
            year = extract_year_from_filename(pdf_path.name)
            if year is None:
                if include_undated:
                    print(f"  [WARN] no year found in filename, including anyway: {pdf_path.name}")
                    found.append(DiscoveredFile(ticker, pdf_path, year))
                else:
                    skipped.append((pdf_path, "no year in filename"))
                continue
            if year >= min_year:
                found.append(DiscoveredFile(ticker, pdf_path, year))
            else:
                skipped.append((pdf_path, f"year {year} < {min_year}"))

    print(f"\nDiscovered {len(found)} document(s) to process, skipped {len(skipped)}:")
    for path, reason in skipped:
        print(f"  [SKIP] {path.name}  ({reason})")
    return found


# =============================================================================
# PER-TICKER FILE SELECTION — keep only actual financial-statement filings,
# capped at a max count per ticker (cheap classifier-model call on filenames
# only, no PDF content needed for this step).
# =============================================================================

FILE_SELECTION_SCHEMA = {
    "type": "object",
    "properties": {
        "selected_files": {
            "type": "array",
            "items": {"type": "string"},
            "description": "Exact filenames (copied verbatim from the list given) "
                            "of the best sources for full financial statements.",
        },
    },
    "required": ["selected_files"],
    "additionalProperties": False,
}

FILE_SELECTION_PROMPT_TEMPLATE = """Ticker: {ticker}

Here are every filing found for this company:
{listing}

Select at most {max_files} of these filenames that are the BEST sources for
extracting full financial statements (income statement / statement of
comprehensive income, balance sheet / statement of financial position, and
cash flow statement).

Strongly prefer:
- Annual financial statements (often tagged AFS, or described as
  "reviewed"/"audited" financial statements)
- Integrated or annual reports (often tagged IAR) IF no separate AFS exists
  for that year, since they normally reproduce the full statements

Do NOT select (unless nothing better exists for a given year):
- AGM notices, proxies, minutes, or explanatory notes
- Debt/bond prospectuses or programme documents
- Remuneration reports, sustainability/ESG reports
- Trading statements, SENS announcements, or other short updates that don't
  contain full statements

If multiple years are present, prefer covering distinct years over
duplicating the same year, up to the cap.

Return ONLY exact filenames copied verbatim from the list above."""


def select_financial_statement_files(client, ticker: str, candidates: list["DiscoveredFile"],
                                      max_files: int = 2) -> list["DiscoveredFile"]:
    """Uses a cheap LLM call over just the filenames (no PDF content) to keep
    only the filings that actually contain full financial statements for
    this ticker - dropping AGM notices, debt prospectuses, remuneration
    reports, etc. - and caps the result at max_files. Always calls the model
    (even when there are few candidates) since count alone doesn't tell us
    whether a filing is actually a statement."""
    if not candidates:
        return []

    filenames = [c.path.name for c in candidates]
    listing = "\n".join(f"- {name}" for name in filenames)
    prompt = FILE_SELECTION_PROMPT_TEMPLATE.format(ticker=ticker, listing=listing,
                                                     max_files=max_files)

    def _fallback(reason: str) -> list["DiscoveredFile"]:
        print(f"  [WARN] {ticker}: {reason}; falling back to most recent {max_files} filing(s) by year")
        return sorted(candidates, key=lambda c: (c.year or 0), reverse=True)[:max_files]

    try:
        response = client.messages.create(
            model=CLASSIFIER_MODEL_NAME,
            max_tokens=1000,
            messages=[{"role": "user", "content": prompt}],
            output_config={"format": {"type": "json_schema", "schema": FILE_SELECTION_SCHEMA}},
        )
    except Exception as e:
        return _fallback(f"file-selection API call failed ({e})")

    text_blocks = [b.text for b in response.content if getattr(b, "type", None) == "text"]
    if not text_blocks:
        return _fallback("file-selection call returned no text block")

    try:
        result = json.loads("".join(text_blocks))
    except json.JSONDecodeError:
        return _fallback("could not parse file-selection response as JSON")

    selected_names = set(result.get("selected_files", []))
    valid_names = set(filenames)
    unknown_names = selected_names - valid_names
    if unknown_names:
        print(f"  [WARN] {ticker}: model selected unrecognized filename(s), ignoring: {sorted(unknown_names)}")
        selected_names &= valid_names

    if not selected_names:
        return _fallback("file-selection returned no valid matches")

    selected = [c for c in candidates if c.path.name in selected_names][:max_files]

    dropped = [name for name in filenames if name not in selected_names]
    for name in dropped:
        print(f"  [SKIP] {name}  (not selected as a core financial-statement filing)")

    return selected


# =============================================================================
# THE ENGINEERED EXTRACTION PROMPT (schema-locked via structured outputs)
# =============================================================================

EXTRACTION_SYSTEM_PROMPT = r"""
You are a financial-statement extraction engine. You will be given the text
(or an image) of ONE page from a company filing. Your only job is to find
structured financial data on that page and return JSON matching the schema
you were given. You are a transcriber, not an analyst: never compute,
estimate, or infer a number that is not printed on the page.

RULES
1. If the page has no financial statement/table, set
   page_contains_financial_statement to false and leave line_items empty.
2. Numbers: strip thousands separators and currency symbols; keep the sign.
   A number in parentheses, e.g. "(1 804)", is NEGATIVE: -1804. An em-dash
   "—" or "*" (less than 1 unit) is 0, not null; null means no value/column
   printed for that line item at all.
3. Do not invent subtotals. Extract "Total assets" etc. only if printed;
   never compute a subtotal yourself.
4. canonical_label: use the closest of the following, else
   "OTHER: <short description>":
   Income statement: Revenue, Cost of Sales, Gross Profit, Operating Expenses,
     Depreciation and Amortisation, Impairment Charges, Finance Income,
     Finance Costs, Net Foreign Exchange Gain/Loss, Share of Associates/JV
     Results, Profit Before Tax, Income Tax Expense, Profit After Tax,
     Profit Attributable to Owners, Profit Attributable to Non-controlling
     Interests, Basic EPS, Diluted EPS
   Balance sheet: Property Plant and Equipment, Intangible Assets and
     Goodwill, Right-of-Use Assets, Investments, Investment in
     Associates/JVs, Deferred Tax Assets, Total Non-current Assets,
     Inventories, Trade and Other Receivables, Cash and Cash Equivalents,
     Total Current Assets, Total Assets, Share Capital, Retained Earnings,
     Other Reserves, Non-controlling Interests, Total Equity, Borrowings
     (Non-current), Lease Liabilities (Non-current), Total Non-current
     Liabilities, Trade and Other Payables, Borrowings (Current),
     Total Current Liabilities, Total Liabilities, Total Equity and
     Liabilities
   Cash flow: Cash Generated From Operations, Interest Received/Paid,
     Income Tax Paid, Net Cash from Operating Activities, Capital
     Expenditure, Net Cash used in Investing Activities, Proceeds/Repayment
     of Borrowings, Dividends Paid, Net Cash used in Financing Activities,
     Net Increase/Decrease in Cash, Cash at Beginning of Period, Cash at
     End of Period
5. Every line item must be traceable to the exact page it came from - do
   not merge figures across pages.
6. If a page shows both consolidated (Group) and standalone (Company)
   columns, extract each under its own entity_name - never blend them.
7. Flag, don't fix: if a value looks like a probable OCR misread, set
   confidence "low" and explain why in flag_reason - never silently guess.
"""

EXTRACTION_JSON_SCHEMA = {
    "type": "object",
    "properties": {
        "page_contains_financial_statement": {"type": "boolean"},
        "statement_type": {
            "anyOf": [
                {
                    "type": "string",
                    "enum": ["income_statement", "balance_sheet", "cash_flow_statement",
                             "equity_statement", "segment_note", "kpi_table",
                             "remuneration_table", "other_financial_table"],
                },
                {"type": "null"},
            ],
        },
        "entity_name": {"type": ["string", "null"]},
        "currency": {"type": ["string", "null"]},
        "unit_scale": {"type": ["string", "null"]},
        "periods": {"type": "array", "items": {"type": "string"}},
        "line_items": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "source_label": {"type": "string"},
                    "canonical_label": {"type": "string"},
                    "note_ref": {"type": ["string", "null"]},
                    "is_subtotal_or_total": {"type": "boolean"},
                    # A fixed-shape list of {period, value} pairs, NOT a dict
                    # keyed by period name. Anthropic's structured outputs
                    # can't validate objects with arbitrary/dynamic property
                    # names (every object needs a known, fixed set of keys),
                    # so "one column per period" has to be an array instead.
                    "values": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "period": {"type": "string"},
                                "value": {"type": ["number", "null"]},
                            },
                            "required": ["period", "value"],
                            "additionalProperties": False,
                        },
                    },
                    "confidence": {"type": "string", "enum": ["high", "medium", "low"]},
                    "flag_reason": {"type": ["string", "null"]},
                },
                "required": ["source_label", "canonical_label", "note_ref",
                             "is_subtotal_or_total", "values", "confidence", "flag_reason"],
                "additionalProperties": False,
            },
        },
    },
    "required": ["page_contains_financial_statement", "statement_type", "entity_name",
                 "currency", "unit_scale", "periods", "line_items"],
    "additionalProperties": False,
}

# Cheap pre-filter so we don't send every page of a 150-page filing to the
# model. A page must contain one of these before it's worth a full extraction
# call. This is a recall-favoring heuristic (false positives are fine and
# cheap to reject; false negatives silently lose data) - tune as needed.
STATEMENT_KEYWORDS = [
    "statement of comprehensive income", "statement of financial position",
    "statement of cash flows", "income statement", "balance sheet",
    "cash flow statement",
]


def page_looks_relevant(text: str) -> bool:
    low = text.lower()
    return any(kw in low for kw in STATEMENT_KEYWORDS)


def page_text_is_reliable(text: str) -> bool:
    """Detects broken-font-encoding pages (glyph codes like '(cid:15)'
    instead of real characters) or near-empty text layers, which need the
    vision fallback instead of the text path."""
    if len(text.strip()) < 30:
        return False
    if text.count("(cid:") > 5:
        return False
    return True


# =============================================================================
# EXTRACTION CALLS
# =============================================================================

def call_llm(client, document_name: str, page_number: int,
             page_text: Optional[str] = None, image_bytes: Optional[bytes] = None,
             model: str = MODEL_NAME) -> dict:
    if image_bytes is not None:
        content = [
            {"type": "image", "source": {"type": "base64", "media_type": "image/png",
                                          "data": base64.standard_b64encode(image_bytes).decode()}},
            {"type": "text", "text": f"Document: {document_name}\nPage: {page_number}"},
        ]
    else:
        content = [{"type": "text", "text":
                    f"Document: {document_name}\nPage: {page_number}\n\n{page_text}"}]

    response = client.messages.create(
        model=model,
        max_tokens=4000,
        system=EXTRACTION_SYSTEM_PROMPT,
        messages=[{"role": "user", "content": content}],
        output_config={"format": {"type": "json_schema", "schema": EXTRACTION_JSON_SCHEMA}},
    )
    text_blocks = [block.text for block in response.content
                   if getattr(block, "type", None) == "text"]
    if not text_blocks:
        raise ValueError(
            f"No text block in response for {document_name} p{page_number} "
            f"(got block types: {[getattr(b, 'type', None) for b in response.content]})"
        )
    return json.loads("".join(text_blocks))


@dataclass
class ExtractedRow:
    ticker: str
    entity_name: str
    statement_type: str
    source_label: str
    canonical_label: str
    note_ref: Optional[str]
    is_subtotal: bool
    period: str
    value: Optional[float]
    currency: str
    unit_scale: str
    confidence: str
    flag_reason: Optional[str]
    document_name: str
    page_number: int


def flatten(raw: dict, ticker: str, document_name: str, page_number: int) -> list[ExtractedRow]:
    """document_name/page_number are passed in from our own Python loop
    (we already know them) rather than asked of the model - one less thing
    that can be wrong or missing in the model's response."""
    rows = []
    if not raw.get("page_contains_financial_statement"):
        return rows
    for li in raw.get("line_items", []):
        for pv in li["values"]:
            rows.append(ExtractedRow(
                ticker=ticker,
                entity_name=raw.get("entity_name") or "",
                statement_type=raw.get("statement_type") or "",
                source_label=li["source_label"],
                canonical_label=li["canonical_label"],
                note_ref=li.get("note_ref"),
                is_subtotal=li.get("is_subtotal_or_total", False),
                period=pv["period"],
                value=pv["value"],
                currency=raw.get("currency") or "",
                unit_scale=raw.get("unit_scale") or "",
                confidence=li.get("confidence", "medium"),
                flag_reason=li.get("flag_reason"),
                document_name=document_name,
                page_number=page_number,
            ))
    return rows


def process_pdf(client, doc: DiscoveredFile) -> list[ExtractedRow]:
    print(f"\nProcessing {doc.ticker}: {doc.path.name} (year={doc.year})")
    rows = []
    with pdfplumber.open(doc.path) as pdf:
        for i, page in enumerate(pdf.pages):
            text = page.extract_text() or ""

            if not page_text_is_reliable(text):
                # Broken font encoding or scanned page - only worth the
                # vision call if it's plausibly a statement page; cheap
                # signal here is weak, so we still try but keep it targeted
                # by page position isn't reliable across companies, so we
                # fall back to rendering + a vision classification pass.
                img = page.to_image(resolution=150).original
                buf = io.BytesIO()
                img.save(buf, format="PNG")
                raw = call_llm(client, doc.path.name, i + 1, image_bytes=buf.getvalue())
            elif page_looks_relevant(text):
                raw = call_llm(client, doc.path.name, i + 1, page_text=text)
            else:
                continue

            if raw.get("page_contains_financial_statement"):
                print(f"  page {i+1}: extracted {raw.get('statement_type')} "
                      f"({len(raw.get('line_items', []))} line items)")
                rows.extend(flatten(raw, doc.ticker, doc.path.name, i + 1))
    return rows


# =============================================================================
# EXCEL EXPORT
# =============================================================================

def write_workbook(rows: list[ExtractedRow], out_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extracted Financials"

    headers = ["Ticker", "Entity", "Statement Type", "Canonical Label", "Source Label",
               "Note", "Is Subtotal", "Period", "Value", "Currency", "Unit Scale",
               "Confidence", "Flag Reason", "Source Document", "Source Page"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F1F1F")
        c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.freeze_panes = "A2"

    for r, row in enumerate(rows, start=2):
        vals = [row.ticker, row.entity_name, row.statement_type, row.canonical_label,
                row.source_label, row.note_ref, row.is_subtotal, row.period, row.value,
                row.currency, row.unit_scale, row.confidence, row.flag_reason,
                row.document_name, row.page_number]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            if row.confidence == "low":
                cell.fill = PatternFill("solid", fgColor="FFE6E6")  # flag low-confidence rows

    for i, w in enumerate([10, 22, 16, 30, 34, 8, 10, 16, 14, 10, 10, 10, 26, 34, 10], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    wb.save(out_path)
    print(f"\nSaved {len(rows)} extracted rows to {out_path}")


# =============================================================================
# CSV EXPORT
# =============================================================================

CSV_HEADERS = ["Ticker", "Entity", "Statement Type", "Canonical Label", "Source Label",
               "Note", "Is Subtotal", "Period", "Value", "Currency", "Unit Scale",
               "Confidence", "Flag Reason", "Source Document", "Source Page"]


def write_csv(rows: list[ExtractedRow], out_path: Path):
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(CSV_HEADERS)
        for row in rows:
            writer.writerow([row.ticker, row.entity_name, row.statement_type,
                              row.canonical_label, row.source_label, row.note_ref,
                              row.is_subtotal, row.period, row.value, row.currency,
                              row.unit_scale, row.confidence, row.flag_reason,
                              row.document_name, row.page_number])
    print(f"Saved {len(rows)} extracted rows to {out_path}")


# =============================================================================
# CORE PIPELINE — call this directly from a notebook cell (no argparse/sys.argv)
# =============================================================================

def _build_output_dir(out_root, tickers: list[str]) -> Path:
    """Builds a ticker-stamped output folder under out_root, e.g.
    'extraction_output_AAPL_MTN' for two tickers, or
    'extraction_output_AAPL_+4more' when there are more than a handful -
    keeps the folder name readable instead of arbitrarily long."""
    uniq = sorted(set(tickers))
    if len(uniq) <= 4:
        stamp = "_".join(uniq)
    else:
        stamp = "_".join(uniq[:3]) + f"_+{len(uniq) - 3}more"
    out_dir = Path(out_root) / f"extraction_output_{stamp}"
    out_dir.mkdir(parents=True, exist_ok=True)
    return out_dir


def run_extraction(root_folder, min_year: int = 2022, skip_undated: bool = False,
                    out: str = "financial_extraction_results.xlsx",
                    out_format: str = "xlsx",
                    out_dir=None) -> list[ExtractedRow]:
    """Runs the full pipeline against real Python arguments — no argparse, no
    sys.argv. This is what a Jupyter notebook cell should call directly:

        rows = run_extraction("/path/to/root_folder", min_year=2022)
        rows = run_extraction("/path/to/root_folder", out="results.csv", out_format="csv")

    out_format: "xlsx" (default), "csv", or "both" (writes both files, using
    `out` as the stem for whichever extension(s) it doesn't already match).

    A ticker-stamped output folder is created automatically (e.g.
    'extraction_output_AAPL_MTN/') and the workbook/csv is written inside
    it - you don't need to create a folder yourself. `out_dir` controls
    *where that folder is created* (defaults to the current directory);
    pass a specific parent path if you want it created elsewhere.

    File selection: after discovery, a cheap classifier-model call (on
    filenames only) picks the best financial-statement filing (AFS/IAR-type
    document) for EACH distinct year found per ticker - i.e. one file per
    year, spanning from min_year through the most recent year present for
    that ticker - dropping AGM notices, debt prospectuses, remuneration
    reports, etc. There's no fixed per-ticker cap to choose; the number of
    files selected simply follows however many distinct years that ticker
    has on disk.

    Returns the list of ExtractedRow (also written to disk per out_format).
    """
    root_folder = Path(root_folder)
    out = Path(out)
    out_format = out_format.lower()
    if out_format not in ("xlsx", "csv", "both"):
        raise ValueError(f"out_format must be 'xlsx', 'csv', or 'both' (got {out_format!r})")

    if not root_folder.is_dir():
        raise NotADirectoryError(f"Not a folder: {root_folder}")

    docs = discover_documents(root_folder, min_year, include_undated=not skip_undated)
    if not docs:
        print("No qualifying PDFs found.")
        return []

    client = get_client()  # prompts for the key here, right before it's first needed

    by_ticker: dict[str, list[DiscoveredFile]] = {}
    for d in docs:
        by_ticker.setdefault(d.ticker, []).append(d)

    print("\nSelecting one financial-statement filing per year for each ticker...")
    docs = []
    for ticker, files in by_ticker.items():
        years_present = sorted({d.year for d in files if d.year is not None})
        # One file per distinct year found (min_year..max year for this ticker).
        # Undated files (year is None) don't add to the year count, but still
        # fall back to len(files) so they aren't silently dropped entirely.
        n_years = len(years_present) if years_present else len(files)
        docs.extend(select_financial_statement_files(client, ticker, files, n_years))

    if not docs:
        print("No files remained after financial-statement selection.")
        return []

    all_rows: list[ExtractedRow] = []
    for doc in docs:
        try:
            all_rows.extend(process_pdf(client, doc))
        except Exception as e:
            print(f"  [ERROR] {doc.path.name}: {e}")

    tickers_processed = sorted({doc.ticker for doc in docs})
    output_dir = _build_output_dir(out_dir if out_dir is not None else Path.cwd(),
                                    tickers_processed)
    out_path = output_dir / out.name
    print(f"\nWriting output to: {output_dir}/")

    if out_format in ("xlsx", "both"):
        write_workbook(all_rows, out_path.with_suffix(".xlsx"))
    if out_format in ("csv", "both"):
        write_csv(all_rows, out_path.with_suffix(".csv"))
    return all_rows


# =============================================================================
# CLI ENTRY POINT — only used when run as `python financial_extraction_master.py ...`
# from an actual terminal. In Jupyter, sys.argv holds the kernel's own launch
# arguments (e.g. "-f kernel-xxx.json"), so argparse would fail there with
# SystemExit: 2 — that's why notebooks should call run_extraction() above
# directly instead of main().
# =============================================================================

def main():
    parser = argparse.ArgumentParser(description="Extract financial statements from a folder of ticker-named PDF filings.")
    parser.add_argument("root_folder", type=Path, help="Path to the folder containing one subfolder per ticker")
    parser.add_argument("--min-year", type=int, default=2022, help="Skip filings older than this (default: 2022)")
    parser.add_argument("--skip-undated", action="store_true", help="Skip files where no year could be parsed from the filename (default: include them, with a warning)")
    parser.add_argument("--out", type=Path, default=Path("financial_extraction_results.xlsx"))
    parser.add_argument("--out-format", choices=["xlsx", "csv", "both"], default="xlsx",
                         help="Output format(s) to write (default: xlsx)")
    parser.add_argument("--out-dir", type=Path, default=None,
                         help="Parent directory in which to create the ticker-stamped "
                              "output folder (default: current directory)")
    args = parser.parse_args()

    try:
        run_extraction(args.root_folder, args.min_year, args.skip_undated, args.out,
                        args.out_format, args.out_dir)
    except NotADirectoryError as e:
        sys.exit(str(e))


# Only parses sys.argv / calls main() when run as a plain script from a
# terminal. Importing this file (e.g. inside a notebook) never triggers it.
if __name__ == "__main__" and "ipykernel" not in sys.modules:
    main()